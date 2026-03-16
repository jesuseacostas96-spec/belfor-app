import streamlit as st
import zipfile, io, re, json, os
from collections import OrderedDict, defaultdict
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="BELFOR Equipment Log", page_icon="🏗️", layout="centered")
st.markdown("""
<style>
.block-container{padding-top:2rem;max-width:760px}
.stButton>button{background:#0D2B55;color:white;border-radius:6px;border:none;padding:0.6rem 2rem;font-weight:700;font-size:15px;width:100%}
</style>""", unsafe_allow_html=True)
st.markdown("""
<div style='background:#0D2B55;padding:1.5rem 2rem;border-radius:10px;margin-bottom:1.5rem'>
<div style='color:#D4A017;font-size:0.75rem;letter-spacing:3px;font-weight:700'>BELFOR PROPERTY RESTORATION</div>
<div style='color:white;font-size:1.6rem;font-weight:800;margin-top:4px'>Equipment Log Generator</div>
<div style='color:#90b4d4;font-size:0.85rem;margin-top:4px'>WhatsApp chat → BELFOR Weekly EQ Template + Summary</div>
</div>""", unsafe_allow_html=True)

try:
    api_key = st.secrets["ANTHROPIC_API_KEY"]
except Exception:
    st.error("API key not found. Add ANTHROPIC_API_KEY to Streamlit Secrets.")
    st.stop()

with st.expander("📱 How to export WhatsApp chat"):
    st.markdown("1. Open WhatsApp group\n2. Tap group name\n3. Export Chat → Without Media\n4. Upload the .zip here")

uploaded   = st.file_uploader("Upload WhatsApp .zip or .txt", type=["zip","txt"], label_visibility="collapsed")
job_name   = st.text_input("Job name / address", placeholder="e.g. Emera Port Royale - 3333 Port Royale Dr")
job_number = st.text_input("Job # / JDE", placeholder="e.g. JDE 100623171")

def extract_txt(f):
    if f.name.endswith(".zip"):
        with zipfile.ZipFile(io.BytesIO(f.read())) as z:
            for name in z.namelist():
                if name.endswith(".txt"):
                    return z.read(name).decode("utf-8", errors="ignore")
    return f.read().decode("utf-8", errors="ignore")

def parse_messages(raw):
    lines = raw.replace("\r","").split("\n")
    messages, current = [], None
    for line in lines:
        m = re.match(r'^\[(\d{1,2}/\d{1,2}/\d{2,4}),\s*(\d{1,2}:\d{2}:\d{2}\s*[AP]M)\]\s*([^:]+):\s*(.+)$', line)
        if m:
            if current: messages.append(current)
            current = {"date":m.group(1),"time":m.group(2),"sender":m.group(3).strip(),"text":m.group(4).strip()}
        elif current and line.strip() and not line.startswith("‎"):
            current["text"] += "\n" + line.strip()
    if current: messages.append(current)
    return messages

def extract_equipment(messages):
    import anthropic
    client = anthropic.Anthropic(api_key=api_key)
    SYSTEM = """Extract equipment placement AND removal from WhatsApp restoration crew messages.

For PLACED equipment (techs saying they installed/placed equipment):
{"date":"3/15/26","unit":"Unit 702","action":"placed","ams":["CODE1"],"dhs":["CODE2"]}

For REMOVED equipment (words like: removed, pulled, saqué, recogí, out, pickup, retiré, moved out, picked up, equipment removed):
{"date":"3/15/26","unit":"Unit 702","action":"removed","ams":["CODE1"],"dhs":[]}

Rules:
- Only confirmed actions, not requests or assessments
- action must be exactly "placed" or "removed"
- Return ONLY a valid JSON array, no explanation, no markdown"""

    all_results = []
    for i in range(0, len(messages), 50):
        chunk = messages[i:i+50]
        text = "\n".join(f"[{m['date']} {m['time']}] {m['sender']}: {m['text']}" for m in chunk)
        try:
            resp = client.messages.create(
                model="claude-sonnet-4-5",
                max_tokens=4000,
                system=SYSTEM,
                messages=[{"role":"user","content":f"Extract:\n\n{text}"}]
            )
            raw = resp.content[0].text.strip().replace("```json","").replace("```","")
            all_results.extend(json.loads(raw))
        except Exception as e:
            st.warning(f"Chunk error: {e}")

    # Consolidate by (date, unit, action)
    consolidated = OrderedDict()
    for r in all_results:
        action = r.get("action","placed")
        key = (r.get("date",""), r.get("unit",""), action)
        if key not in consolidated:
            consolidated[key] = {"date":r["date"],"unit":r["unit"],"action":action,"ams":[],"dhs":[]}
        consolidated[key]["ams"].extend(r.get("ams",[]))
        consolidated[key]["dhs"].extend(r.get("dhs",[]))
    return list(consolidated.values())

def get_floor(unit):
    u = unit.lower()
    for k,v in [("7th",7),("6th",6),("5th",5),("4th",4),("3rd",3),("2nd",2),("1st",1)]:
        if k in u: return v
    m = re.search(r"(\d+)\s*(?:st|nd|rd|th)?\s*floor", u)
    if m: return int(m.group(1))
    try: return int(''.join(filter(str.isdigit,unit))) // 100
    except: return 0

FLOOR_LABELS = {7:"7th Floor",6:"6th Floor",5:"5th Floor",4:"4th Floor",
                3:"3rd Floor",2:"2nd Floor",1:"1st Floor"}

def sort_unit(unit):
    if "hallway" in unit.lower(): return (0, unit)
    try: return (1, -int(''.join(filter(str.isdigit,unit))))
    except: return (1, unit)

def build_excel(equipment, job_name, job_number):
    AM_BG  = "EBF4FB"; DH_BG  = "FEF9E7"
    REM_BG = "FFE5E5"  # light red for removed
    thin   = Side(style="thin",   color="CCCCCC")
    thick  = Side(style="medium", color="0D2B55")

    template_path = os.path.join(os.path.dirname(__file__), "Weekly_EQ.xlsx")
    wb = load_workbook(template_path)

    # Group by floor (placed only for floor tabs, but track removed too)
    floors_placed  = defaultdict(list)
    floors_removed = defaultdict(list)
    all_dates = sorted(set(e["date"] for e in equipment))

    for loc in equipment:
        f = get_floor(loc["unit"])
        if loc.get("action","placed") == "placed":
            floors_placed[f].append(loc)
        else:
            floors_removed[f].append(loc)

    # Combine floors from both placed and removed
    all_floors = sorted(set(list(floors_placed.keys()) + list(floors_removed.keys())), reverse=True)

    first_floor_num = all_floors[0]
    wb.active.title = FLOOR_LABELS.get(first_floor_num, f"Floor {first_floor_num}")

    # ── FLOOR TABS ──────────────────────────────────────────────────────
    for i, floor_num in enumerate(all_floors):
        label = FLOOR_LABELS.get(floor_num, f"Floor {floor_num}")
        if i == 0:
            ws = wb.active
        else:
            src = wb[FLOOR_LABELS.get(first_floor_num, f"Floor {first_floor_num}")]
            ws = wb.copy_worksheet(src)
            ws.title = label

        ws["A2"] = job_number or ""
        ws["E2"] = job_name or ""

        # Date columns: I=9, K=11, M=13, O=15, Q=17, S=19, U=21
        date_cols = [9, 11, 13, 15, 17, 19, 21]
        date_col_map = {}
        for di, d in enumerate(all_dates[:7]):
            col = date_cols[di]
            ws.cell(row=10, column=col).value = d
            ws.cell(row=11, column=col).value = None
            ws.cell(row=12, column=col).value = None
            date_col_map[d] = col

        # Build rows: placed first, then removed (for this floor)
        placed  = sorted(floors_placed.get(floor_num, []),  key=lambda x: sort_unit(x["unit"]))
        removed = sorted(floors_removed.get(floor_num, []), key=lambda x: sort_unit(x["unit"]))

        all_rows = []
        for loc in placed:
            first = True
            for code in loc["ams"]:
                all_rows.append({"code":code,"desc":"AM","location":loc["unit"] if first else "",
                                  "date":loc["date"],"action":"placed","first":first})
                first = False
            for code in loc["dhs"]:
                all_rows.append({"code":code,"desc":"DH","location":loc["unit"] if first else "",
                                  "date":loc["date"],"action":"placed","first":first})
                first = False
        for loc in removed:
            first_r = True
            for code in loc["ams"]:
                all_rows.append({"code":code,"desc":"AM","location":loc["unit"] if first_r else "",
                                  "date":loc["date"],"action":"removed","first":first_r})
                first_r = False
            for code in loc["dhs"]:
                all_rows.append({"code":code,"desc":"DH","location":loc["unit"] if first_r else "",
                                  "date":loc["date"],"action":"removed","first":first_r})
                first_r = False

        # Clear rows 13-62
        for r in range(13, 63):
            for c in [3,5,7,9,11,13,15,17,19,21]:
                ws.cell(row=r, column=c).value = None

        for idx, eq in enumerate(all_rows[:50]):
            row = 13 + idx
            is_am     = eq["desc"] == "AM"
            is_removed = eq["action"] == "removed"
            status    = 6 if is_removed else 2
            bg = REM_BG if is_removed else (AM_BG if is_am else DH_BG)
            txt_color = "C0392B" if is_removed else ("1A4F7A" if is_am else "7B5200")

            for col, val in [(3,eq["code"]),(5,eq["desc"]),(7,eq["location"])]:
                c = ws.cell(row=row, column=col, value=val)
                c.fill = PatternFill("solid", start_color=bg)
                c.font = Font(name="Arial", size=9, bold=(col==5), color=txt_color)
                c.alignment = Alignment(horizontal="left", vertical="center")
                c.border = Border(left=thin,right=thin,
                                  top=(thick if eq["first"] else thin), bottom=thin)

            # Status in correct date column
            if eq["date"] in date_col_map:
                dcol = date_col_map[eq["date"]]
                c = ws.cell(row=row, column=dcol, value=status)
                c.fill = PatternFill("solid", start_color=bg)
                c.font = Font(name="Arial", size=9, bold=True, color=txt_color)
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = Border(left=thin,right=thin,
                                  top=(thick if eq["first"] else thin), bottom=thin)

        am_placed  = sum(1 for r in all_rows if r["desc"]=="AM" and r["action"]=="placed")
        dh_placed  = sum(1 for r in all_rows if r["desc"]=="DH" and r["action"]=="placed")
        am_removed = sum(1 for r in all_rows if r["desc"]=="AM" and r["action"]=="removed")
        dh_removed = sum(1 for r in all_rows if r["desc"]=="DH" and r["action"]=="removed")
        ws.cell(row=63, column=9).value = am_placed - am_removed
        ws.cell(row=65, column=9).value = dh_placed - dh_removed

    # ── SUMMARY TAB ─────────────────────────────────────────────────────
    ws_sum = wb.create_sheet("📋 Summary")
    ws_sum.sheet_view.showGridLines = False

    NAVY  = "0D2B55"; BLUE = "1560A4"; GOLD = "D4A017"; WHITE = "FFFFFF"
    AM_H  = "1A4F7A"; DH_H = "7B5200"
    GREEN = "E2EFDA"; RED  = "FFE5E5"; LGREY = "F5F7FA"
    med   = Side(style="medium", color=NAVY)
    thn   = Side(style="thin",   color="CCCCCC")
    def tb(c): c.border = Border(left=med,right=med,top=med,bottom=med)
    def nb(c): c.border = Border(left=thn,right=thn,top=thn,bottom=thn)
    def hset(c, val, bg, fg=WHITE, sz=10, bold=True, center=True):
        c.value = val
        c.font  = Font(name="Arial", bold=bold, color=fg, size=sz)
        c.fill  = PatternFill("solid", start_color=bg)
        c.alignment = Alignment(horizontal="center" if center else "left",
                                vertical="center", wrap_text=False)

    # Col widths
    for col, w in [(1,5),(2,22),(3,12),(4,12),(5,12),(6,12),(7,12),(8,12),(9,12)]:
        ws_sum.column_dimensions[get_column_letter(col)].width = w

    # Title
    ws_sum.merge_cells("A1:I1")
    hset(ws_sum["A1"], "BELFOR PROPERTY RESTORATION  —  EQUIPMENT SUMMARY", NAVY, GOLD, sz=13)
    tb(ws_sum["A1"]); ws_sum.row_dimensions[1].height = 28

    ws_sum.merge_cells("A2:I2")
    hset(ws_sum["A2"], f"{job_name or ''}   |   {job_number or ''}   |   Date(s): {', '.join(all_dates)}", BLUE, WHITE, sz=10, bold=False)
    tb(ws_sum["A2"]); ws_sum.row_dimensions[2].height = 18

    # ── SECTION 1: Affected locations ──
    ws_sum.merge_cells("A4:I4")
    hset(ws_sum["A4"], "AFFECTED LOCATIONS", NAVY, GOLD, sz=11)
    tb(ws_sum["A4"]); ws_sum.row_dimensions[4].height = 22

    loc_headers = ["#", "Unit / Location", "Type", "AMs Placed", "DHs Placed", "AMs Removed", "DHs Removed", "Net AMs", "Net DHs"]
    for ci, h in enumerate(loc_headers, 1):
        c = ws_sum.cell(row=5, column=ci, value=h)
        bg = AM_H if "AM" in h else (DH_H if "DH" in h else BLUE)
        hset(c, h, bg, WHITE, sz=9); tb(c)
    ws_sum.row_dimensions[5].height = 20

    # Build per-location summary
    loc_summary = OrderedDict()
    for e in equipment:
        key = e["unit"]
        if key not in loc_summary:
            loc_summary[key] = {"unit":key,"type":"Hallway" if "hallway" in key.lower() else "Apartment",
                                 "am_placed":0,"dh_placed":0,"am_removed":0,"dh_removed":0}
        if e.get("action","placed") == "placed":
            loc_summary[key]["am_placed"] += len(e["ams"])
            loc_summary[key]["dh_placed"] += len(e["dhs"])
        else:
            loc_summary[key]["am_removed"] += len(e["ams"])
            loc_summary[key]["dh_removed"] += len(e["dhs"])

    # Sort descending
    def loc_sort(unit):
        if "hallway" in unit.lower():
            for k,v in [("7th",7),("6th",6),("5th",5),("4th",4),("3rd",3),("2nd",2),("1st",1)]:
                if k in unit.lower(): return (1,-v,unit)
            return (1,0,unit)
        try: return (0,-int(''.join(filter(str.isdigit,unit))),unit)
        except: return (0,0,unit)

    sorted_locs = sorted(loc_summary.values(), key=lambda x: loc_sort(x["unit"]))
    apt_count  = sum(1 for l in sorted_locs if l["type"]=="Apartment")
    hall_count = sum(1 for l in sorted_locs if l["type"]=="Hallway")

    row = 6
    for idx, loc in enumerate(sorted_locs):
        bg = LGREY if idx % 2 == 0 else WHITE
        net_am = loc["am_placed"] - loc["am_removed"]
        net_dh = loc["dh_placed"] - loc["dh_removed"]
        vals = [idx+1, loc["unit"], loc["type"],
                loc["am_placed"], loc["dh_placed"],
                loc["am_removed"] or "—", loc["dh_removed"] or "—",
                net_am, net_dh]
        for ci, val in enumerate(vals, 1):
            c = ws_sum.cell(row=row, column=ci, value=val)
            row_bg = RED if (ci in (6,7) and loc["am_removed"]+loc["dh_removed"] > 0) else bg
            c.fill = PatternFill("solid", start_color=row_bg)
            c.font = Font(name="Arial", size=9,
                          bold=(ci in (8,9)),
                          color=("C0392B" if ci in (6,7) and loc["am_removed"]+loc["dh_removed"]>0
                                 else "1A4F7A" if ci == 8 else "7B5200" if ci == 9 else "000000"))
            c.alignment = Alignment(horizontal="center" if ci != 2 else "left", vertical="center")
            nb(c)
        ws_sum.row_dimensions[row].height = 16
        row += 1

    # Location totals row
    tot_amp = sum(l["am_placed"]  for l in sorted_locs)
    tot_dhp = sum(l["dh_placed"]  for l in sorted_locs)
    tot_amr = sum(l["am_removed"] for l in sorted_locs)
    tot_dhr = sum(l["dh_removed"] for l in sorted_locs)

    ws_sum.merge_cells(f"A{row}:C{row}")
    hset(ws_sum.cell(row,1), f"TOTALS  —  {apt_count} Apts  |  {hall_count} Hallways  |  {len(sorted_locs)} Locations",
         NAVY, GOLD, sz=10); tb(ws_sum.cell(row,1))
    for ci, val in [(4,tot_amp),(5,tot_dhp),(6,tot_amr),(7,tot_dhr),(8,tot_amp-tot_amr),(9,tot_dhp-tot_dhr)]:
        bg = AM_H if ci in (4,6,8) else DH_H
        hset(ws_sum.cell(row,ci), val, bg, GOLD, sz=11); tb(ws_sum.cell(row,ci))
    ws_sum.row_dimensions[row].height = 24
    row += 2

    # ── SECTION 2: Daily summary ──
    ws_sum.merge_cells(f"A{row}:I{row}")
    hset(ws_sum.cell(row,1), "DAILY EQUIPMENT SUMMARY", NAVY, GOLD, sz=11)
    tb(ws_sum.cell(row,1)); ws_sum.row_dimensions[row].height = 22
    row += 1

    # Headers: Type | Date1 | Date2 ... | TOTAL
    day_headers = ["Metric"] + all_dates + ["CUMULATIVE"]
    for ci, h in enumerate(day_headers, 1):
        c = ws_sum.cell(row=row, column=ci)
        hset(c, h, BLUE, WHITE, sz=9); tb(c)
        ws_sum.column_dimensions[get_column_letter(ci)].width = 14 if ci > 1 else 22
    ws_sum.row_dimensions[row].height = 20
    row += 1

    # Calculate per-day stats
    def day_stats(date, action, desc):
        return sum(len(e[desc+"s"]) for e in equipment
                   if e["date"]==date and e.get("action","placed")==action)

    metrics = [
        ("AMs Placed",   "placed",  "am",  AM_BG,  "1A4F7A"),
        ("AMs Removed",  "removed", "am",  RED,    "C0392B"),
        ("Net AMs",      "net",     "am",  GREEN,  "1A6B1A"),
        ("DHs Placed",   "placed",  "dh",  DH_BG,  "7B5200"),
        ("DHs Removed",  "removed", "dh",  RED,    "C0392B"),
        ("Net DHs",      "net",     "dh",  GREEN,  "1A6B1A"),
    ]

    for label, action, typ, bg, color in metrics:
        c = ws_sum.cell(row=row, column=1, value=label)
        c.font = Font(name="Arial", bold=True, size=9, color=color)
        c.fill = PatternFill("solid", start_color=bg)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        nb(c)

        cumulative = 0
        for di, date in enumerate(all_dates):
            col = di + 2
            if action == "net":
                placed_val  = sum(len(e[typ+"s"]) for e in equipment
                                  if e["date"]==date and e.get("action","placed")=="placed")
                removed_val = sum(len(e[typ+"s"]) for e in equipment
                                  if e["date"]==date and e.get("action","placed")=="removed")
                val = placed_val - removed_val
            else:
                val = sum(len(e[typ+"s"]) for e in equipment
                          if e["date"]==date and e.get("action","placed")==action)
            cumulative += val
            c = ws_sum.cell(row=row, column=col, value=val if val else "—")
            c.font = Font(name="Arial", bold=True, size=10, color=color)
            c.fill = PatternFill("solid", start_color=bg)
            c.alignment = Alignment(horizontal="center", vertical="center")
            nb(c)

        # Cumulative total
        c = ws_sum.cell(row=row, column=len(all_dates)+2, value=cumulative)
        c.font = Font(name="Arial", bold=True, size=11, color=color)
        c.fill = PatternFill("solid", start_color=NAVY if "Net" in label else bg)
        if "Net" in label:
            c.font = Font(name="Arial", bold=True, size=11, color=GOLD)
        c.alignment = Alignment(horizontal="center", vertical="center")
        tb(c)
        ws_sum.row_dimensions[row].height = 20
        row += 1

    # ── LEGEND ──
    row += 1
    ws_sum.merge_cells(f"A{row}:I{row}")
    hset(ws_sum.cell(row,1),
         "STATUS LEGEND:  2 = Equipment Running   |   6 = Equipment Removed from Job   |   Red rows = Removed equipment",
         "FFF3CD", "7B5200", sz=9, bold=False)
    tb(ws_sum.cell(row,1)); ws_sum.row_dimensions[row].height = 18

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)

    total_am_net = tot_amp - tot_amr
    total_dh_net = tot_dhp - tot_dhr
    return buf, tot_amp, tot_dhp, tot_amr, tot_dhr, total_am_net, total_dh_net, apt_count, hall_count

if uploaded:
    st.markdown("### Generate BELFOR Weekly EQ Template")
    if st.button("⚙️ Process & Generate Excel"):
        with st.spinner("Reading chat..."):
            raw = extract_txt(uploaded)
            messages = parse_messages(raw)
        st.info(f"Found {len(messages)} messages")

        with st.spinner("Analyzing with AI (placed + removed)..."):
            equipment = extract_equipment(messages)

        placed_count  = sum(1 for e in equipment if e.get("action","placed")=="placed")
        removed_count = sum(1 for e in equipment if e.get("action","placed")=="removed")
        if removed_count > 0:
            st.info(f"✅ {placed_count} placement records  |  🔴 {removed_count} removal records detected")

        if not equipment:
            st.error("No equipment records found.")
        else:
            with st.spinner("Building Excel..."):
                try:
                    buf, tot_amp, tot_dhp, tot_amr, tot_dhr, net_am, net_dh, apt_c, hall_c = build_excel(
                        equipment, job_name, job_number)
                    st.success("✅ Done!")
                    c1,c2,c3,c4,c5 = st.columns(5)
                    c1.metric("Locations", apt_c + hall_c)
                    c2.metric("AMs Placed", tot_amp)
                    c3.metric("DHs Placed", tot_dhp)
                    c4.metric("Net AMs", net_am, delta=f"-{tot_amr} removed" if tot_amr else None)
                    c5.metric("Net DHs", net_dh, delta=f"-{tot_dhr} removed" if tot_dhr else None)
                    fname = (job_number or "Equipment").replace(" ","_") + "_Weekly_EQ.xlsx"
                    st.download_button(
                        "⬇️ Download BELFOR Weekly EQ",
                        buf, fname,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True)
                except Exception as e:
                    st.error(f"Error: {e}")

st.markdown("---")
st.markdown("<div style='text-align:center;color:#999;font-size:0.75rem'>BELFOR Property Restoration • Powered by Claude AI</div>", unsafe_allow_html=True)
